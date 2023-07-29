import openpyxl
file_path = "20221231.xlsx"
Sheet1Name = 'Sheet1'
FilePath2 = 'Hubspot Import Test 1.xlsx'
Sheet2Name = 'Test 1'


def GetRecords():
    storerecords = list()
    StoreDeal = list()
    book = openpyxl.load_workbook(file_path)
    user_data = book.get_sheet_by_name(str(Sheet1Name))
    for x in range(1,user_data.max_row):
        CadgeCode = str(user_data[x][1].value)
        ExpirationDate = str(user_data[x][2].value)
        DealName = str(user_data[x][3].value)
        StoreDeal.append(DealName)
        storerecords.append({"CadgeCode":CadgeCode,"ExpirationDate":ExpirationDate,"DealName":DealName})
    return storerecords
def TestHubspot():
    storerecords = list()
    book = openpyxl.load_workbook(FilePath2)
    user_data = book.get_sheet_by_name(str(Sheet2Name))
    for x in range(1,user_data.max_row):
        DealName = str(user_data[x][0].value)
        storerecords.append(DealName)
    return storerecords

def matchResult():
    DataHubspot = TestHubspot()
    Records = GetRecords()
    RemoveDuplicates = list()
    for i in Records:
        GetDeals = i.get('DealName')
        ExpirationDate = i.get('ExpirationDate')
        CadgeCode = i.get('CadgeCode')
        total = GetDeals,ExpirationDate,CadgeCode
        if GetDeals in DataHubspot:
            RemoveDuplicates.append({"Deal":GetDeals,"ExpirationDate":i.get('ExpirationDate'),"CadgeCode":i.get('CadgeCode')})
    return RemoveDuplicates


def writeSheet():
    data= matchResult()
    from openpyxl import Workbook
    book = Workbook()
    sheet = book.active
    for i in data:
        deal = i.get('Deals')
        ExpirationDate = i.get('ExpirationDate')
        CadgeCode = i.get('CadgeCode')
        rows = (
            (deal, ExpirationDate, CadgeCode),
        )
        for row in rows:
            sheet.append(row)
        book.save('newFile.xlsx')
writeSheet()